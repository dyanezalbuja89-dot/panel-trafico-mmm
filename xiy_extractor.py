#!/usr/bin/env python3
"""
Extrae las campañas de Ford 2026 desde Xiy (xiy.today) usando el Next.js Flight payload
embebido en el HTML inicial (no requiere browser headless, no requiere login).

Input:
  - Google Sheet "CONSOLIDADO CAMPAÑAS FORD" (4 pestañas: ENERO/FEBRERO/MARZO/ABRIL)
    URL: https://docs.google.com/spreadsheets/d/19yig-cc6UoOKi8OxRJEkARyLIWkPkP8kcvkfXOSnL3Y/

Output:
  - data_xiy.json: estructura por mes → campañas → líneas con inversión, CPA, conversiones

Uso:
  python3 xiy_extractor.py
"""

import json
import re
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl

SHEET_ID = "19yig-cc6UoOKi8OxRJEkARyLIWkPkP8kcvkfXOSnL3Y"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
OUTPUT = Path(__file__).parent / "data_xiy.json"
TMP_XLSX = Path("/tmp/ford_campaigns.xlsx")

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0 Safari/537.36")


def download_sheet():
    print(f"[1/3] Descargando sheet → {TMP_XLSX}")
    req = urllib.request.Request(SHEET_URL, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=30) as resp:
        TMP_XLSX.write_bytes(resp.read())


def read_campaigns_index():
    """Lee las 4 pestañas (ENERO/FEBRERO/MARZO/ABRIL) y devuelve lista de
    {month, campaign, url} para los flujos."""
    wb = openpyxl.load_workbook(TMP_XLSX, data_only=True)
    out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        # encontrar header
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, 8)]
            if any(v and "CAMPAÑA" in str(v).upper() for v in vals):
                header_row = r
                break
        if header_row is None:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            camp = ws.cell(r, 2).value
            url = ws.cell(r, 3).value
            if camp and url and "xiy.today" in str(url):
                out.append({
                    "month": sname.capitalize(),
                    "campaign": str(camp).strip(),
                    "url": str(url).strip(),
                })
    print(f"[2/3] Encontradas {len(out)} campañas en el sheet")
    return out


def fetch_html(url, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", errors="ignore")


def extract_flight_payload(html):
    """Next.js mete los props en chunks de self.__next_f.push([1,"..."]).
    Los concatena y devuelve el string crudo."""
    matches = re.findall(r'self\.__next_f\.push\(\[1,"(.+?)"\]\)', html, re.DOTALL)
    # Cada match es un string escapado con JS. Decodificarlo:
    chunks = []
    for m in matches:
        # En el HTML viene como JSON-string escapado dentro de un argumento JS.
        # Lo decodificamos pasándolo por json.loads de "...":
        try:
            chunk = json.loads('"' + m + '"')
        except json.JSONDecodeError:
            chunk = m
        chunks.append(chunk)
    return "".join(chunks)


def parse_flujo(html):
    """Extrae los datos del flujo: suborder (cabecera) + lines (filas de inversión)."""
    payload = extract_flight_payload(html)

    result = {
        "raw_size": len(payload),
        "suborder": None,
        "order": None,
        "lines": [],
    }

    # ORDER (el flujo mensual completo, p.ej. "OM0020705-Normal-Ford-Enero-2026")
    # Estructura: tiene "code":20705,"month":"Enero","brand":"Ford","totalAmount":<X>
    # Buscamos un objeto con "brand":"Ford" y "month": presente
    for m in re.finditer(r'\{[^{}]*"brand":"Ford"[^{}]*\}', payload):
        block = m.group(0)
        if '"month"' in block and '"totalAmount"' in block:
            try:
                result["order"] = json.loads(block)
                break
            except json.JSONDecodeError:
                pass

    # SUBORDER: tiene "campaign":"...","totalAmount":..., código de campaña, etc.
    # Lo identifica que NO tiene "brand":"Ford" pero SÍ tiene "campaign"
    for m in re.finditer(r'\{[^{}]*"campaign":"[^"]+"[^{}]*\}', payload):
        block = m.group(0)
        if '"totalAmount"' in block and '"brand"' not in block:
            try:
                result["suborder"] = json.loads(block)
                break
            except json.JSONDecodeError:
                pass

    # LINES: array de líneas. Buscar "lines":[ ... ] balanceado.
    i = payload.find('"lines":[')
    if i >= 0:
        # parse manual con balance de corchetes
        start = i + len('"lines":')
        depth = 0
        end = None
        for j in range(start, len(payload)):
            c = payload[j]
            if c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    end = j + 1
                    break
        if end:
            lines_str = payload[start:end]
            try:
                # las "lines" pueden contener "$D2026-..." (fechas serializadas Next.js).
                # Eliminamos el prefijo $D antes de parsear:
                lines_str_clean = re.sub(r'"\$D(\d{4}-\d{2}-\d{2}T[^"]*)"', r'"\1"', lines_str)
                # Eliminar referencias tipo "$NN" (Next.js refs)
                lines_str_clean = re.sub(r'"\$\d+"', 'null', lines_str_clean)
                # Eliminar "$undefined"
                lines_str_clean = lines_str_clean.replace('"$undefined"', 'null')
                result["lines"] = json.loads(lines_str_clean)
            except json.JSONDecodeError as e:
                result["lines_parse_error"] = str(e)
                result["lines_raw"] = lines_str[:500]

    return result


def classify_modelo_from_format(fmt):
    """Mapea el campo `format` de una LÍNEA Xiy a un modelo Ford normalizado.
    OJO: en Xiy el `format` a veces es el modelo Ford (RANGER XLT, F150, etc.)
    y a veces es el formato publicitario (TOP FEED, PAGE POST, DEMAND GEN, etc.).
    Esta función SOLO devuelve modelo si el format claramente identifica uno;
    si no, devuelve None y se intenta fallback por nombre de campaña."""
    if not fmt:
        return None
    f = str(fmt).upper().strip()
    # Normalizar typos conocidos
    f = f.replace("TERRITOTY", "TERRITORY").replace("RANGERL", "RANGER")
    # Detectar primero si es formato publicitario (no modelo)
    AD_FORMATS = ("TOP FEED", "PAGE POST", "PAGEPOST", "DEMAND GEN", "SEARCH KEYWORDS",
                  "UNIVERSO", "COBERT", "CLICS ESTIMADOS", "REACH ESTIMADO",
                  "VIDEO ADS", "CAROUSEL", "STORIES")
    if any(af in f for af in AD_FORMATS):
        return None  # no es modelo, es ad format
    # Modelos Ford
    if "RANGER" in f:
        if "XLT" in f: return "Ranger XLT"
        if "EXONERADO" in f: return "Ranger XL Exonerados"
        if "XL" in f:  return "Ranger XL"
        return "Ranger"
    if "F150" in f or "F-150" in f or "F 150" in f: return "F-150"
    if "EVEREST" in f: return "Everest"
    if "TERRITORY" in f: return "Territory"
    if "ESCAPE" in f:
        if "ST LINE" in f or "STLINE" in f or "ST-LINE" in f: return "Escape ST Line"
        if "1.5" in f: return "Escape 1.5"
        return "Escape"
    if "EXPLORER" in f: return "Explorer"
    if "MAVERICK" in f: return "Maverick"
    if "BRONCO" in f: return "Bronco"
    if "MUSTANG" in f: return "Mustang"
    if "TRANSIT" in f: return "Transit"
    if "EDGE" in f: return "Edge"
    return None  # No es modelo conocido


def classify_modelo_from_campaign(campaign_name):
    """FALLBACK: si la línea no tiene format, usar el nombre de la campaña.
    Devuelve None si la campaña es genérica (AYF, POSICIONAMIENTO, etc.)."""
    n = (campaign_name or "").upper()
    if "RANGER" in n:
        if "XLT" in n: return "Ranger XLT"
        if "EXONERADO" in n: return "Ranger XL Exonerados"
        if " XL" in n:  return "Ranger XL"
        return "Ranger"
    if "F150" in n or "F-150" in n: return "F-150"
    if "EVEREST" in n: return "Everest"
    if "TERRITORY" in n: return "Territory"
    if "ESCAPE" in n:
        if "ST LINE" in n or "STLINE" in n: return "Escape ST Line"
        return "Escape"
    return None  # genérico — la línea tendrá que decir el modelo


def categorize_campaign(campaign_name):
    """Devuelve un TIPO de iniciativa (Performance/Awareness/Activación/Marca)
    basado en el nombre de la campaña."""
    n = (campaign_name or "").upper()
    if "LEADS" in n: return "Performance - Leads"
    if "AYF" in n: return "Awareness - AYF Regional"
    if "POSICIONAMIENTO" in n: return "Awareness - Posicionamiento Producto"
    if "LANZAMIENTO" in n: return "Lanzamiento"
    if "OPEN HOUSE" in n: return "Activación - Open House"
    if "RACE WEEKEND" in n: return "Activación - Race Weekend"
    if "POWERDAYS" in n: return "Activación - PowerDays"
    if "BLINDADOS" in n: return "Producto - Blindados"
    if "SEGUIDORES" in n: return "Marca - Seguidores"
    if "RENTING" in n: return "Servicios - Renting"
    if "RENOVATION" in n: return "Marca - Renovation"
    if "UTILIDADES" in n: return "Promo - Utilidades"
    if "BRANDING" in n: return "Awareness - Branding"
    if "INTERACCIÓN" in n or "INTERACCION" in n: return "Awareness - Interacción"
    return "Otros"


def extract_region_from_audience(audience):
    """Extrae la región (Sierra/Costa/Manabí/Nacional) desde el audience text."""
    if not audience: return "Sin clasificar"
    a = audience.upper()
    # ORGU: Sierra = solo Quito; Costa = Guayaquil/Manta/Cuenca; Manabí aparte
    if "SIERRA" in a or "QUITO" in a or "TUMBACO" in a or "LA Y" in a: return "Sierra (Quito)"
    if "COSTA" in a or "GUAYAQUIL" in a: return "Costa"
    if "MANABI" in a or "MANABÍ" in a or "MANTA" in a: return "Manabí"
    if "CJA" in a or "CUENCA" in a or "ORELLANA" in a: return "Cuenca/Orellana"
    if "MACHALA" in a: return "Machala"
    if "EMPRESAS" in a: return "B2B/Empresas"
    if "EC " in a or "ECUADOR" in a or "CIUDADES FORD" in a: return "Nacional"
    return "Otra"


def enrich_line(line):
    """Parsea el campo additionalData (JSON-string) y extrae CPA, Conversiones."""
    ad = line.get("additionalData")
    if isinstance(ad, str):
        try:
            adp = json.loads(ad)
        except json.JSONDecodeError:
            adp = {}
    elif isinstance(ad, dict):
        adp = ad
    else:
        adp = {}
    # claves típicas: "CPA", "CONVERSIONES"
    line["cpa"] = adp.get("CPA") or adp.get("cpa")
    line["conversiones_esperadas"] = adp.get("CONVERSIONES") or adp.get("Conversiones") or adp.get("conversiones")
    try:
        line["cpa"] = float(line["cpa"]) if line["cpa"] is not None else None
    except (ValueError, TypeError):
        pass
    try:
        line["conversiones_esperadas"] = float(line["conversiones_esperadas"]) if line["conversiones_esperadas"] is not None else None
    except (ValueError, TypeError):
        pass
    return line


def process_one(item):
    """Descarga y parsea un flujo. Devuelve dict con metadata + parsed.

    Atribución de modelo:
      1) Cada LÍNEA tiene un campo `format` que es el modelo Ford real.
      2) Si la línea no tiene format, fallback al modelo derivado del nombre de campaña.
      3) Si tampoco eso, queda como 'Genérico'.
    """
    try:
        html = fetch_html(item["url"])
        parsed = parse_flujo(html)
        lines = [enrich_line(l) for l in parsed.get("lines", [])]
        sub = parsed.get("suborder") or {}

        # Fallback model por campaña (para líneas sin format)
        fallback_modelo = classify_modelo_from_campaign(item["campaign"])

        # Categoría de la campaña (para clasificar líneas sin modelo)
        cat = categorize_campaign(item["campaign"])
        # Etiqueta sustituta cuando la línea es awareness multi-modelo
        if "AYF Regional" in cat:        unattributed_label = "Awareness Regional (AYF)"
        elif "Posicionamiento" in cat:   unattributed_label = "Awareness Posicionamiento"
        elif "Marca" in cat:             unattributed_label = "Marca"
        elif "Activación" in cat:        unattributed_label = "Activación / Eventos"
        elif "Servicios" in cat:         unattributed_label = "Servicios"
        elif "Lanzamiento" in cat:       unattributed_label = "Lanzamiento (Multi-modelo)"
        else:                             unattributed_label = "Sin atribuir"

        # Enriquecer cada línea con modelo y región
        for l in lines:
            line_modelo = classify_modelo_from_format(l.get("format"))
            if line_modelo:
                l["modelo"] = line_modelo
                l["modelo_source"] = "format"
            elif fallback_modelo:
                l["modelo"] = fallback_modelo
                l["modelo_source"] = "campaign_name"
            else:
                l["modelo"] = unattributed_label
                l["modelo_source"] = "categoria"
            l["region"] = extract_region_from_audience(l.get("audience"))

        # Fallback: si no se pudo parsear suborder, calcular total desde lines
        total_amount = sub.get("totalAmount")
        total_investment = sub.get("totalInvestment")
        if total_amount is None and lines:
            total_amount = sum((l.get("amount") or 0) for l in lines)
        if total_investment is None and lines:
            total_investment = sum((l.get("investment") or 0) for l in lines)
        total_conversiones = sum((l.get("conversiones_esperadas") or 0) for l in lines)

        return {
            **item,
            "categoria": categorize_campaign(item["campaign"]),
            "modelo_principal": fallback_modelo or "Multi-modelo",
            "ok": True,
            "totalAmount": total_amount,
            "totalInvestment": total_investment,
            "totalConversionesEsperadas": total_conversiones,
            "status": sub.get("status"),
            "objective": sub.get("objective"),
            "code": sub.get("code"),
            "lines": lines,
            "lines_count": len(lines),
            "lines_parse_error": parsed.get("lines_parse_error"),
        }
    except Exception as e:
        return {**item, "ok": False, "error": str(e)}


def consolidate_for_panel(flat_lines):
    """Mapea los modelos finos de Xiy a los nombres normalizados del panel
    (DATA.conversion_data.FORD.por_modelo usa keys: RANGER, F-150, EVEREST,
    TERRITORY, ESCAPE, EXPLORER, EXPEDITION, BRONCO).

    Cualquier variante de RANGER/ESCAPE colapsa al modelo padre.
    Lo que no es modelo Ford atribuible (Awareness Regional AYF, Marca,
    Activación/Eventos, Servicios, Sin atribuir, Lanzamiento multi-modelo, etc.)
    se queda fuera del cruce modelo-a-modelo y se devuelve en `non_modelo`.
    """
    # Modelos canónicos que el panel reconoce (uppercase, matchea por_modelo)
    PANEL_MODELS = ["RANGER", "F-150", "EVEREST", "TERRITORY", "ESCAPE",
                    "EXPLORER", "EXPEDITION", "BRONCO"]

    def normalize_modelo(modelo_xiy):
        """Devuelve (canonical_name, is_modelo_ford). Si no es modelo Ford
        atribuible, devuelve (modelo_xiy, False)."""
        if not modelo_xiy:
            return ("Sin atribuir", False)
        m = str(modelo_xiy).upper().strip()
        # RANGER (todas las variantes): RANGER, RANGER XLT, RANGER XL,
        # RANGER XL EXONERADOS
        if "RANGER" in m:
            return ("RANGER", True)
        if "F-150" in m or "F150" in m or "F 150" in m:
            return ("F-150", True)
        if "EVEREST" in m:
            return ("EVEREST", True)
        if "TERRITORY" in m:
            return ("TERRITORY", True)
        # ESCAPE (todas las variantes): ESCAPE, ESCAPE 1.5, ESCAPE ST LINE
        if "ESCAPE" in m:
            return ("ESCAPE", True)
        if "EXPLORER" in m:
            return ("EXPLORER", True)
        if "EXPEDITION" in m:
            return ("EXPEDITION", True)
        if "BRONCO" in m:
            return ("BRONCO", True)
        # NO atribuible a modelo Ford: AYF Regional, Marca, Activación, etc.
        return (modelo_xiy, False)

    # Agencias ORGU (matchea con DATA.conversion_data.FORD.por_agencia)
    PANEL_AGENCIAS = ["Tumbaco", "La Y", "CJA", "Orellana", "Manta",
                      "Machala", "Portoviejo"]

    def map_audience_to_agencias(audience):
        """Mapea el texto de audience a una o más agencias ORGU.
        - Mención directa (TUMBACO, LA Y, CJA, MANTA, MACHALA, ORELLANA) → atribución 100%
          a esa agencia.
        - Mención regional ('Concesionarios Sierra/Costa') → atribución repartida
          entre las agencias de la región.
        - Sin clasificar / Nacional → None (queda en bucket 'Multi/Nacional').
        Devuelve: lista de agencias o None.
        """
        if not audience:
            return None
        a = audience.upper()
        # Mención directa a una agencia específica → 100% a ella
        # (puede haber audiencias multi-ciudad como CJA/ORELLANA/MANTA — devolvemos
        #  todas las que aparezcan para repartir)
        direct = []
        if "TUMBACO" in a:                direct.append("Tumbaco")
        if "LA Y" in a or " LAY " in a:   direct.append("La Y")
        if "CJA" in a or "CUENCA" in a:   direct.append("CJA")
        if "ORELLANA" in a:               direct.append("Orellana")
        if "MANTA" in a:                  direct.append("Manta")
        if "MACHALA" in a:                direct.append("Machala")
        if "PORTOVIEJO" in a:             direct.append("Portoviejo")
        if direct:
            return direct
        # Mención regional sin agencia específica
        if "SIERRA" in a:        return ["Tumbaco", "La Y"]
        if "MANABI" in a or "MANABÍ" in a: return ["Manta", "Portoviejo"]
        if "COSTA" in a:         return ["Machala"]  # ORGU costa = Machala
        # B2B / nacional / sin clasificar
        return None

    months_order = ["Enero", "Febrero", "Marzo", "Abril", "Mayo"]
    consolidated = {
        "months": {},           # {mes: {MODELO: {amount, convers, n_lines}}}
        "totals_modelo": {},    # {MODELO: {amount, convers, n_lines}}
        "totals_mes": {},       # {mes: total_amount}
        "non_modelo": {},       # {label_no_atribuible: {amount, convers, n_lines}}
        "totals_mes_non_modelo": {},  # {mes: total no atribuible} útil para info
        "total_general": 0.0,
        "total_atribuible_modelo": 0.0,
        "total_non_modelo": 0.0,
        "panel_models": PANEL_MODELS,
        # NUEVO: agregación por agencia (mapeada desde audience)
        "totals_agencia": {},   # {agencia: {amount, n_lines}}
        "agencia_modelo": {},   # {agencia: {MODELO: amount}}  cruce agencia×modelo
        "agencia_mes": {},      # {agencia: {mes: amount}}
        "total_multi_nacional": 0.0,  # inversión sin agencia atribuible
        "panel_agencias": PANEL_AGENCIAS,
        # NUEVO: agregación por medio (Meta/Google/TikTok)
        "totals_medio": {},     # {medio: {amount, n_lines}}
        "medio_objective": {},  # {medio: {objective: amount}}
    }

    for L in flat_lines:
        mes = L.get("month")
        if not mes:
            continue
        canonical, is_modelo = normalize_modelo(L.get("modelo"))
        amount = float(L.get("amount") or 0)
        convers = float(L.get("conversiones_esperadas") or 0)

        consolidated["total_general"] += amount
        consolidated["totals_mes"].setdefault(mes, 0.0)
        consolidated["totals_mes"][mes] += amount

        if is_modelo:
            consolidated["total_atribuible_modelo"] += amount
            # months × modelo
            consolidated["months"].setdefault(mes, {})
            consolidated["months"][mes].setdefault(canonical,
                {"amount": 0.0, "convers": 0.0, "n_lines": 0})
            consolidated["months"][mes][canonical]["amount"] += amount
            consolidated["months"][mes][canonical]["convers"] += convers
            consolidated["months"][mes][canonical]["n_lines"] += 1
            # totals modelo
            consolidated["totals_modelo"].setdefault(canonical,
                {"amount": 0.0, "convers": 0.0, "n_lines": 0})
            consolidated["totals_modelo"][canonical]["amount"] += amount
            consolidated["totals_modelo"][canonical]["convers"] += convers
            consolidated["totals_modelo"][canonical]["n_lines"] += 1
        else:
            consolidated["total_non_modelo"] += amount
            consolidated["non_modelo"].setdefault(canonical,
                {"amount": 0.0, "convers": 0.0, "n_lines": 0})
            consolidated["non_modelo"][canonical]["amount"] += amount
            consolidated["non_modelo"][canonical]["convers"] += convers
            consolidated["non_modelo"][canonical]["n_lines"] += 1
            consolidated["totals_mes_non_modelo"].setdefault(mes, 0.0)
            consolidated["totals_mes_non_modelo"][mes] += amount

        # --- Atribución por AGENCIA (basada en audience) ---
        ags = map_audience_to_agencias(L.get("audience"))
        if ags:
            share = amount / len(ags)  # repartir equitativo si menciona varias
            for ag in ags:
                consolidated["totals_agencia"].setdefault(ag,
                    {"amount": 0.0, "n_lines": 0})
                consolidated["totals_agencia"][ag]["amount"] += share
                consolidated["totals_agencia"][ag]["n_lines"] += 1
                # cruce agencia × modelo (sólo para modelos Ford)
                if is_modelo:
                    consolidated["agencia_modelo"].setdefault(ag, {})
                    consolidated["agencia_modelo"][ag].setdefault(canonical, 0.0)
                    consolidated["agencia_modelo"][ag][canonical] += share
                # cruce agencia × mes
                consolidated["agencia_mes"].setdefault(ag, {})
                consolidated["agencia_mes"][ag].setdefault(mes, 0.0)
                consolidated["agencia_mes"][ag][mes] += share
        else:
            consolidated["total_multi_nacional"] += amount

        # --- Atribución por MEDIO (Meta / Google / TikTok) ---
        medio = (L.get("media") or "Sin medio").strip()
        # Normalizar variantes
        if medio.upper() in ("TIK TOK", "TIKTOK", "TIK-TOK"):
            medio = "TikTok"
        elif medio.upper() == "META":
            medio = "Meta"
        elif medio.upper() == "GOOGLE":
            medio = "Google"
        objective = (L.get("objective") or "Sin objetivo").strip()
        consolidated["totals_medio"].setdefault(medio,
            {"amount": 0.0, "n_lines": 0})
        consolidated["totals_medio"][medio]["amount"] += amount
        consolidated["totals_medio"][medio]["n_lines"] += 1
        consolidated["medio_objective"].setdefault(medio, {})
        consolidated["medio_objective"][medio].setdefault(objective, 0.0)
        consolidated["medio_objective"][medio][objective] += amount

    # Redondear amounts a 2 decimales para limpieza en JSON
    def _round_dict(d, key="amount", nd=2):
        for v in d.values():
            if isinstance(v, dict) and key in v:
                v[key] = round(v[key], nd)
                if "convers" in v:
                    v["convers"] = round(v["convers"], 2)

    for mes_d in consolidated["months"].values():
        _round_dict(mes_d)
    _round_dict(consolidated["totals_modelo"])
    _round_dict(consolidated["non_modelo"])
    for k in list(consolidated["totals_mes"]):
        consolidated["totals_mes"][k] = round(consolidated["totals_mes"][k], 2)
    for k in list(consolidated["totals_mes_non_modelo"]):
        consolidated["totals_mes_non_modelo"][k] = round(
            consolidated["totals_mes_non_modelo"][k], 2)
    consolidated["total_general"] = round(consolidated["total_general"], 2)
    consolidated["total_atribuible_modelo"] = round(
        consolidated["total_atribuible_modelo"], 2)
    consolidated["total_non_modelo"] = round(consolidated["total_non_modelo"], 2)
    consolidated["months_order"] = [m for m in months_order
                                    if m in consolidated["totals_mes"]]

    # Redondear agencia / medio
    _round_dict(consolidated["totals_agencia"])
    _round_dict(consolidated["totals_medio"])
    for ag, mods in consolidated["agencia_modelo"].items():
        for k in list(mods):
            mods[k] = round(mods[k], 2)
    for ag, meses in consolidated["agencia_mes"].items():
        for k in list(meses):
            meses[k] = round(meses[k], 2)
    for med, objs in consolidated["medio_objective"].items():
        for k in list(objs):
            objs[k] = round(objs[k], 2)
    consolidated["total_multi_nacional"] = round(
        consolidated["total_multi_nacional"], 2)
    return consolidated


def main():
    if not TMP_XLSX.exists():
        download_sheet()
    items = read_campaigns_index()

    print(f"[3/3] Fetching {len(items)} flujos en paralelo (5 hilos)...")
    t0 = time.time()
    results = []
    with ThreadPoolExecutor(max_workers=5) as pool:
        futs = {pool.submit(process_one, it): it for it in items}
        for i, fut in enumerate(as_completed(futs), 1):
            r = fut.result()
            results.append(r)
            mark = "✓" if r["ok"] else "✗"
            amt = f"USD {r.get('totalAmount') or 0:>10,.2f}" if r["ok"] else r.get("error", "?")
            print(f"  [{i:>2}/{len(items)}] {mark} {r['month']:<8} {r['campaign']:<40} → {amt}")

    print(f"\nTomó {time.time()-t0:.1f}s")

    # Aplanar a líneas para agregaciones precisas (cada línea tiene su modelo real)
    flat_lines = []
    for r in results:
        if not r["ok"]: continue
        for l in r.get("lines", []):
            flat_lines.append({
                "month": r["month"],
                "campaign": r["campaign"],
                "categoria": r["categoria"],
                "modelo": l["modelo"],
                "modelo_source": l["modelo_source"],
                "region": l["region"],
                "audience": l.get("audience"),
                "media": l.get("media"),
                "objective": l.get("objective"),
                "format": l.get("format"),
                "amount": l.get("amount") or 0,
                "investment": l.get("investment") or 0,
                "cpa": l.get("cpa"),
                "conversiones_esperadas": l.get("conversiones_esperadas") or 0,
                "startDate": l.get("startDate"),
                "endDate": l.get("endDate"),
            })

    # Resumen por mes
    print("\n=== RESUMEN POR MES ===")
    by_month = {}
    for L in flat_lines:
        by_month.setdefault(L["month"], {"n_lines": 0, "amount": 0.0, "convers": 0})
        by_month[L["month"]]["n_lines"] += 1
        by_month[L["month"]]["amount"] += L["amount"]
        by_month[L["month"]]["convers"] += L["conversiones_esperadas"]
    order = ["Enero", "Febrero", "Marzo", "Abril", "Mayo"]
    for m in order:
        if m in by_month:
            d = by_month[m]
            print(f"  {m:<10}: {d['n_lines']:>3} líneas · USD {d['amount']:>12,.2f} · {int(d['convers']):>5} convers. esperadas")

    # Resumen por modelo (atribución real, línea por línea)
    print("\n=== RESUMEN POR MODELO (atribución por LÍNEA, todos los meses) ===")
    by_modelo = {}
    for L in flat_lines:
        m = L["modelo"]
        by_modelo.setdefault(m, {"n_lines": 0, "amount": 0.0, "convers": 0, "sources": {}})
        by_modelo[m]["n_lines"] += 1
        by_modelo[m]["amount"] += L["amount"]
        by_modelo[m]["convers"] += L["conversiones_esperadas"]
        s = L["modelo_source"]
        by_modelo[m]["sources"][s] = by_modelo[m]["sources"].get(s, 0) + 1
    for mod, d in sorted(by_modelo.items(), key=lambda x: -x[1]["amount"]):
        srcs = ", ".join(f"{k}={v}" for k,v in d["sources"].items())
        print(f"  {mod:<25}: {d['n_lines']:>3} líneas · USD {d['amount']:>11,.2f} · {int(d['convers']):>5} convers · [{srcs}]")

    # Resumen por categoría
    print("\n=== RESUMEN POR CATEGORÍA DE CAMPAÑA ===")
    by_cat = {}
    for r in results:
        if r["ok"]:
            by_cat.setdefault(r["categoria"], 0)
            by_cat[r["categoria"]] += r["totalAmount"] or 0
    for cat, total in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"  {cat:<40}: USD {total:>11,.2f}")

    # Cross modelo × mes
    print("\n=== INVERSIÓN MODELO × MES (USD) ===")
    cross = {}
    for L in flat_lines:
        cross.setdefault(L["modelo"], {m:0 for m in order})
        cross[L["modelo"]][L["month"]] += L["amount"]
    headers = ["MODELO"] + [m for m in order if m in by_month]
    print("  " + " | ".join(f"{h:<15}" if h=="MODELO" else f"{h:>10}" for h in headers))
    for mod in sorted(cross.keys(), key=lambda m: -sum(cross[m].values())):
        row = cross[mod]
        cells = [f"{mod:<15}"] + [f"{row[m]:>10,.0f}" for m in headers[1:]]
        print("  " + " | ".join(cells))

    # Consolidación lista para el panel (modelos colapsados a nombres del panel)
    consolidated_for_panel = consolidate_for_panel(flat_lines)

    print("\n=== CONSOLIDATED FOR PANEL (modelos canónicos) ===")
    print(f"  Total general: USD {consolidated_for_panel['total_general']:>11,.2f}")
    print(f"  Atribuible a modelo Ford: USD {consolidated_for_panel['total_atribuible_modelo']:>11,.2f}")
    print(f"  NO atribuible a modelo: USD {consolidated_for_panel['total_non_modelo']:>11,.2f}")
    for mod, d in sorted(consolidated_for_panel["totals_modelo"].items(),
                         key=lambda x: -x[1]["amount"]):
        print(f"    {mod:<12}: USD {d['amount']:>10,.2f} · {int(d['convers']):>5} convers")

    # Guardar JSON
    output = {
        "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source": "https://docs.google.com/spreadsheets/d/" + SHEET_ID,
        "n_campaigns": sum(1 for r in results if r["ok"]),
        "n_failed": sum(1 for r in results if not r["ok"]),
        "n_lines": len(flat_lines),
        "campaigns": results,
        "lines_flat": flat_lines,  # para que el panel pueda iterar fácil
        "summary_by_month": by_month,
        "summary_by_modelo": by_modelo,
        "summary_by_categoria": by_cat,
        "cross_modelo_mes": cross,
        "consolidated_for_panel": consolidated_for_panel,
    }
    OUTPUT.write_text(json.dumps(output, indent=2, ensure_ascii=False, default=str))
    print(f"\n💾 Guardado en {OUTPUT}")


if __name__ == "__main__":
    main()
